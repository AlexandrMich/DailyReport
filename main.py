# import pywhatkit
# pywhatkit.sendwhatmsg('+77058124392','Nice work', 11, 57, 10, True, 1)
from openpyxl import Workbook, load_workbook

wb = load_workbook('2021.08.02.xlsx')
ws = wb.active
last_row = ws.max_row

for i in range(11, last_row):
    var = ws.cell(i, 7).value
    var = str(var)

    if var == "None":
        last_row = i
        break

List = [9, 13, 15]

for i in range(11, last_row):
    for j in range(0, 3):
        print(ws.cell(i, List[j]).value)

wb.save('DailyReport.xlsx')